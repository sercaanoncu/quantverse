"""Validate generated QuantVerse v2 release-candidate artifacts.

The validator checks generated evidence files after a local v2 demo run. It
does not download data, modify portfolio logic or commit generated artifacts.
It writes a machine-readable validation summary under data/processed.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from pathlib import Path, PureWindowsPath
from xml.etree import ElementTree

import numpy as np
import pandas as pd
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "src"))

from project.research.global_numerical_integrity import (
    validate_v2_numerical_integrity,
)  # noqa: E402
from project.research.global_visual_analytics import (  # noqa: E402
    validate_visual_analytics_outputs,
)
from project.research.run_identity import (  # noqa: E402
    RUN_FIELDS,
    validate_registered_artifacts,
)
from project.reporting.artifact_publication import (  # noqa: E402
    validate_publication_manifest,
)
from scripts.audit_quantverse_v2_missing_data_operations import (  # noqa: E402
    scan_repository as scan_missing_data_operations,
)

PROCESSED = ROOT / "data" / "processed"
OUTPUT = ROOT / "output"
VALIDATION_PATH = PROCESSED / "quantverse_v2_artifact_validation.json"

REPORT_PUBLICATION_TYPE = "quantverse_v2_pdf_html_research_package"
REPORT_PUBLICATION_ARTIFACTS = (
    "output/pdf/quantverse_v2_executive_research_report.pdf",
    "output/pdf/quantverse_v2_methodology_validation_appendix.pdf",
    "output/html/quantverse_v2_research_report.html",
    "output/pdf/quantverse_v2_research_report.pdf",
)
EXCEL_PUBLICATION_TYPE = "quantverse_v2_analytical_workbook"
EXCEL_PUBLICATION_ARTIFACTS = ("output/excel/quantverse_v2_research_output.xlsx",)

PUBLICATION_SOURCE_ARTIFACTS = (
    "quantverse_v2_reference_math_checks.csv",
    "quantverse_v2_reference_math_summary.json",
    "quantverse_v2_visual_validation.csv",
    "global_portfolio_league.csv",
    "global_model_selection_report.csv",
    "global_portfolio_league_weights.csv",
    "global_top_holdings_explanation.csv",
    "global_portfolio_risk_report.csv",
    "global_walk_forward_model_comparison.csv",
    "global_walk_forward_leakage_audit.csv",
    "global_walk_forward_leakage_audit.csv",
    "global_walk_forward_uncertainty.csv",
    "global_walk_forward_random_benchmark_provenance.json",
    "global_stress_test_results.csv",
    "quantverse_v2_visual_exposure.csv",
    "quantverse_v2_visual_equity_curve.csv",
    "quantverse_v2_visual_drawdown_curve.csv",
    "quantverse_v2_visual_model_risk_return.csv",
    "quantverse_v2_visual_random_benchmark.csv",
    "quantverse_v2_visual_forecast_error.csv",
    "global_risk_metric_sanity_checks.csv",
    "global_security_identity_audit.csv",
    "global_security_history_eligibility.csv",
    "global_final_model_decision.json",
)
ADDITIONAL_PUBLICATION_SOURCE_ARTIFACTS = (
    "data/universe/current_global_equity_universe.csv",
    "data/processed/global_master_equal_weight_comparison.csv",
    "data/processed/global_master_random_portfolio_benchmark.csv",
    "data/processed/global_model_selection_diagnostics.csv",
    "data/processed/global_exact_proxy_classification_report.csv",
)

REQUIRED_JSON_FIELDS = {
    "quantverse_v2_demo_summary.json": [
        "run_status",
        "final_selected_model",
        "final_model_selection_method",
        "final_model_selection_score",
        "final_model_selection_decision",
        "promotion_decision",
        "weight_sum",
        "final_selected_holdings",
        "run_id",
        "execution_id",
        "data_as_of_date",
        "universe_snapshot_id",
        "data_snapshot_id",
        "config_hash",
        "input_fingerprint",
    ],
    "global_final_model_decision.json": [
        "final_selected_model",
        "final_model_selection_method",
        "final_model_selection_score",
        "final_decision",
        "final_decision_reason",
        "publish_readiness_status",
        "run_id",
        "execution_id",
        "data_as_of_date",
        "universe_snapshot_id",
        "data_snapshot_id",
        "config_hash",
        "input_fingerprint",
    ],
    "quantverse_v2_run_manifest.json": [
        "run_id",
        "execution_id",
        "data_as_of_date",
        "generated_at",
        "universe_snapshot_id",
        "data_snapshot_id",
        "config_hash",
        "input_fingerprint",
    ],
    "quantverse_v2_reference_math_summary.json": [
        "status",
        "check_count",
        "failed_check_count",
        "run_id",
        "execution_id",
        "data_as_of_date",
        "generated_at",
        "universe_snapshot_id",
        "data_snapshot_id",
        "config_hash",
        "input_fingerprint",
        "checks_path",
    ],
    "global_walk_forward_random_benchmark_provenance.json": [
        "benchmark_scope",
        "provenance_status",
        "protocol_hash",
        "fold_schedule_hash",
        "selected_universe_by_fold_hash",
        "model_oos_dates_hash",
        "random_oos_dates_hash",
        "random_weights_hash",
        "oos_dates_match",
        "run_id",
        "config_hash",
        "input_fingerprint",
        "universe_snapshot_id",
        "data_snapshot_id",
    ],
}

REQUIRED_CSVS = [
    "global_model_selection_report.csv",
    "global_model_selection_diagnostics.csv",
    "global_portfolio_league.csv",
    "global_portfolio_league_weights.csv",
    "global_portfolio_risk_report.csv",
    "global_stress_test_results.csv",
    "global_risk_metric_sanity_checks.csv",
    "global_fx_prices.csv",
    "global_walk_forward_model_comparison.csv",
    "global_walk_forward_random_distribution.csv",
    "global_walk_forward_random_returns.csv",
    "global_walk_forward_random_weights.csv",
    "global_walk_forward_uncertainty.csv",
    "global_random_portfolio_percentile_report.csv",
    "global_robustness_sensitivity.csv",
    "global_top_holdings_explanation.csv",
    "global_selected_stocks_report_view.csv",
    "global_selected_stocks_report_view_quality.csv",
    "global_security_identity_audit.csv",
    "global_security_history_eligibility.csv",
    "global_feature_history_eligibility.csv",
    "global_cross_artifact_count_reconciliation.csv",
    "quantverse_v2_artifact_run_registry.csv",
    "global_forecast_validation_by_horizon.csv",
    "global_listing_country_exposure.csv",
    "global_issuer_country_exposure.csv",
    "global_economic_country_exposure.csv",
    "global_industry_exposure.csv",
    "global_exchange_exposure.csv",
    "global_exposure_metadata_quality.csv",
    "quantverse_v2_visual_analytics_summary.csv",
    "quantverse_v2_visual_equity_curve.csv",
    "quantverse_v2_visual_drawdown_curve.csv",
    "quantverse_v2_visual_model_risk_return.csv",
    "quantverse_v2_visual_forecast_error.csv",
    "quantverse_v2_visual_random_benchmark.csv",
    "quantverse_v2_visual_exposure.csv",
    "quantverse_v2_visual_top_holdings.csv",
    "quantverse_v2_visual_validation.csv",
    "quantverse_v2_reference_math_checks.csv",
    "quantverse_v2_missing_data_operation_audit.csv",
]

REQUIRED_HTML_SECTIONS = [
    "Executive Summary",
    "Stock Scoring",
    "Portfolio Model League",
    "Robust Model Selection",
    "Walk-Forward",
    "Exposure",
    "Visual Portfolio Analytics",
    "Equity Curve and Drawdown",
    "Model Risk-Return Map",
    "Forecast Error Versus Random Walk",
    "Random Benchmark Distribution",
    "Exposure and Concentration",
    "Security Identity and History Eligibility",
    "Limitations",
]

REQUIRED_EXCEL_SHEETS = [
    "EXECUTIVE_DASHBOARD",
    "PORTFOLIO",
    "HOLDINGS_DETAIL",
    "MODEL_COMPARISON",
    "MODEL_DECISIONS",
    "UNCERTAINTY",
    "RISK",
    "EXPOSURE",
    "FORECASTS",
    "ELIGIBILITY",
    "AUDIT_FINDINGS",
    "DECISION_REGISTER",
    "FORMULA_DICTIONARY",
    "DATA_DICTIONARY",
    "PORTFOLIO_DASHBOARD",
    "VISUAL_ANALYTICS_DASHBOARD",
    "START_HERE",
    "EXECUTIVE_SUMMARY",
    "SELECTED_STOCKS",
    "SELECTED_STOCKS_RAW",
    "SELECTED_METADATA_QUALITY",
    "SECURITY_IDENTITY",
    "HISTORY_ELIGIBILITY",
    "FEATURE_ELIGIBILITY",
    "COUNT_RECONCILIATION",
    "STOCK_SCORES",
    "RETURN_FORECASTS",
    "MODEL_LEAGUE",
    "MODEL_SELECTION",
    "MODEL_SELECTION_DIAGNOSTICS",
    "FINAL_MODEL_DECISION",
    "FINAL_WEIGHTS",
    "RISK_METRICS",
    "RISK_CONTRIBUTIONS",
    "WALK_FORWARD",
    "RANDOM_PERCENTILES",
    "ROBUSTNESS",
    "EXPOSURE_REGION",
    "EXPOSURE_COUNTRY",
    "EXPOSURE_LISTING_COUNTRY",
    "EXPOSURE_ISSUER_COUNTRY",
    "EXPOSURE_ECON_COUNTRY",
    "EXPOSURE_CURRENCY",
    "EXPOSURE_EXCHANGE",
    "EXPOSURE_INDUSTRY",
    "EXPOSURE_METADATA",
    "TOP_HOLDINGS_EXPLANATION",
    "FORECAST_VALIDATION",
    "VISUAL_SUMMARY",
    "VISUAL_EQUITY_CURVE",
    "VISUAL_DRAWDOWN",
    "VISUAL_RISK_RETURN",
    "VISUAL_FORECAST_ERROR",
    "VISUAL_RANDOM_BENCH",
    "VISUAL_EXPOSURE",
    "VISUAL_TOP_HOLDINGS",
    "VISUAL_VALIDATION",
    "WARNINGS",
    "CLAIM_CONTROL",
]

FORBIDDEN_REPORT_PATTERNS = [
    r"\bguaranteed alpha\b",
    r"\bguaranteed outperformance\b",
    r"\bofficial exact top-100 is supported\b",
    r"\bofficial exact top-100 supported\b",
    r"\binstitutional point-in-time backtest completed\b",
    r"\binstitutional pit backtest completed\b",
    r"\bis investment advice\b",
    r"\bprovides investment advice\b",
    r"\bproduction trading system\b",
    r"\blive trading system\b",
    r"\bbuy recommendation\b",
    r"\bsell recommendation\b",
]

STALE_CURRENT_DECISION_PATTERNS = [
    r"Final model set to Equal Weight",
    r"best metric candidate Min CVaR was not used",
    r"max_crypto_ok",
    r"selected Equal Weight as the public-data research final model",
]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", default=str(ROOT))
    args = parser.parse_args()
    root = Path(args.root).resolve()
    result = validate_artifacts(root)
    processed = root / "data" / "processed"
    processed.mkdir(parents=True, exist_ok=True)
    output_path = processed / "quantverse_v2_artifact_validation.json"
    output_path.write_text(json.dumps(result, indent=2, default=str), encoding="utf-8")
    print(f"artifact_validation_status={result['overall_status']}")
    print(f"artifact_validation_path={output_path}")
    return 0 if result["overall_status"] == "passed" else 1


def validate_artifacts(root: Path) -> dict[str, object]:
    """Validate generated v2 artifacts below ``root``."""
    processed = root / "data" / "processed"
    checks: list[dict[str, object]] = []
    summary = _read_json(processed / "quantverse_v2_demo_summary.json")
    decision = _read_json(processed / "global_final_model_decision.json")
    league = _read_csv(processed / "global_portfolio_league.csv")
    selection = _read_csv(processed / "global_model_selection_report.csv")
    weights = _read_csv(processed / "global_portfolio_league_weights.csv")

    _check_required_json_fields(processed, checks)
    _check_required_csvs(processed, checks)
    _check_demo_run_status(summary, checks)
    _check_final_model_consistency(summary, decision, league, selection, checks)
    _check_final_weights(summary, weights, checks)
    _check_pdf(root / "output" / "pdf" / "quantverse_v2_research_report.pdf", checks)
    _check_pdf(
        root / "output" / "pdf" / "quantverse_v2_executive_research_report.pdf",
        checks,
        label="executive_research_pdf",
    )
    _check_pdf(
        root / "output" / "pdf" / "quantverse_v2_methodology_validation_appendix.pdf",
        checks,
        label="methodology_validation_pdf",
    )
    _check_html(root / "output" / "html" / "quantverse_v2_research_report.html", checks)
    _check_excel(
        root / "output" / "excel" / "quantverse_v2_research_output.xlsx", checks
    )
    _check_excel_formula_errors(
        root / "output" / "excel" / "quantverse_v2_research_output.xlsx",
        checks,
    )
    _check_publication_manifests(root, summary, checks)
    _check_report_claim_language(root, checks)
    _check_numerical_integrity(root, summary, checks)
    _check_reference_math(processed, checks)
    _check_model_selection_evidence(processed, checks)
    _check_publication_source_registry(root, processed, checks)
    _check_missing_data_operations(processed, checks)
    _check_visual_analytics(root, checks)
    _check_exposure_metadata_quality(processed, checks)
    _check_security_identity_history(processed, checks)
    checks.extend(validate_selected_stock_report_semantics(root)["checks"])
    _check_current_v2_reports_no_stale_decisions(root, summary, checks)

    failed = [check for check in checks if not check["passed"]]
    return {
        "overall_status": "passed" if not failed else "failed",
        "check_count": len(checks),
        "failed_check_count": len(failed),
        "checks": checks,
        "summary": {
            "final_selected_model": summary.get("final_selected_model"),
            "final_model_decision": decision.get("final_decision"),
            "promotion_decision": summary.get("promotion_decision"),
            "publish_readiness_status": decision.get("publish_readiness_status"),
        },
    }


def _check_required_json_fields(
    processed: Path, checks: list[dict[str, object]]
) -> None:
    for filename, fields in REQUIRED_JSON_FIELDS.items():
        path = processed / filename
        payload = _read_json(path)
        missing = [field for field in fields if field not in payload]
        checks.append(
            _check(
                f"json_schema_{filename}",
                path.exists() and not missing,
                f"missing_fields={missing}",
            )
        )


def _check_required_csvs(processed: Path, checks: list[dict[str, object]]) -> None:
    for filename in REQUIRED_CSVS:
        path = processed / filename
        frame = _read_csv(path)
        checks.append(
            _check(
                f"csv_non_empty_{filename}",
                path.exists() and not frame.empty,
                f"rows={len(frame)}",
            )
        )


def _check_missing_data_operations(
    processed: Path, checks: list[dict[str, object]]
) -> None:
    """Fail closed when active missing-data operations lack reviewed controls."""
    audit = _read_csv(processed / "quantverse_v2_missing_data_operation_audit.csv")
    required = {
        "operation_id",
        "path",
        "line",
        "function",
        "operation",
        "callsite_fingerprint",
        "source_tree_hash",
        "code",
        "classification",
        "risk_level",
        "status",
        "approved",
        "reason",
        "required_control",
    }
    schema_ok = bool(not audit.empty and required.issubset(audit.columns))
    approved = audit["approved"].map(_truthy) if schema_ok else pd.Series(dtype=bool)
    operations = (
        audit["operation"].astype(str).str.strip().str.lower()
        if schema_ok
        else pd.Series(dtype=str)
    )
    classifications = (
        audit["classification"].astype(str) if schema_ok else pd.Series(dtype=str)
    )
    code = audit["code"].astype(str).str.lower() if schema_ok else pd.Series(dtype=str)
    ffill_rows = operations.eq("ffill")
    ffill_bounded = bool(
        not ffill_rows.any()
        or (
            classifications.loc[ffill_rows].eq("BOUNDED_FORWARD_FILL").all()
            and code.loc[ffill_rows].str.contains("limit=", regex=False).all()
        )
    )
    no_backward_fill = bool(schema_ok and not operations.eq("bfill").any())
    root_path = processed.parents[1]
    fresh_match = True
    fresh_details = "source_rescan_not_available"
    if (root_path / "src").is_dir() and (root_path / "scripts").is_dir():
        fresh = scan_missing_data_operations(root_path)
        comparison_columns = [
            "path",
            "line",
            "function",
            "operation",
            "callsite_fingerprint",
            "classification",
            "approved",
            "source_tree_hash",
        ]
        fresh_match = bool(
            schema_ok
            and set(comparison_columns).issubset(fresh.columns)
            and audit[comparison_columns]
            .sort_values(comparison_columns[:-1], kind="stable")
            .reset_index(drop=True)
            .astype(str)
            .equals(
                fresh[comparison_columns]
                .sort_values(comparison_columns[:-1], kind="stable")
                .reset_index(drop=True)
                .astype(str)
            )
        )
        fresh_details = (
            f"stored_source_hash={audit['source_tree_hash'].iloc[0] if schema_ok else 'missing'}; "
            f"current_source_hash={fresh['source_tree_hash'].iloc[0] if not fresh.empty else 'missing'}"
        )
    all_reviewed = bool(
        schema_ok
        and approved.all()
        and audit["status"].astype(str).eq("reviewed").all()
        and not classifications.eq("REVIEW_REQUIRED_NUMERIC_ZERO_FILL").any()
        and ffill_bounded
        and no_backward_fill
        and fresh_match
    )
    rejected = (
        audit.loc[~approved, "operation_id"].astype(str).head(10).tolist()
        if schema_ok
        else ["missing_or_invalid_schema"]
    )
    checks.append(
        _check(
            "missing_data_operations_are_explicitly_reviewed",
            all_reviewed,
            (
                f"rows={len(audit)}; rejected={rejected}; "
                f"bounded_forward_fill={ffill_bounded}; "
                f"backward_fill_count={int(operations.eq('bfill').sum())}; "
                f"fresh_scan_match={fresh_match}; {fresh_details}"
            ),
        )
    )


def _check_model_selection_evidence(
    processed: Path, checks: list[dict[str, object]]
) -> None:
    """Verify model selection uses comparable walk-forward OOS net evidence."""
    walk = _read_csv(processed / "global_walk_forward_model_comparison.csv")
    selection = _read_csv(processed / "global_model_selection_report.csv")
    randoms = _read_csv(processed / "global_walk_forward_random_distribution.csv")
    random_returns = _read_csv(processed / "global_walk_forward_random_returns.csv")
    random_weights = _read_csv(processed / "global_walk_forward_random_weights.csv")
    model_returns = _read_csv(processed / "global_walk_forward_returns.csv")
    provenance = _read_json(
        processed / "global_walk_forward_random_benchmark_provenance.json"
    )
    manifest = _read_json(processed / "quantverse_v2_run_manifest.json")
    robustness = _read_json(processed / "global_parameter_sensitivity_summary.json")
    uncertainty = _read_csv(processed / "global_walk_forward_uncertainty.csv")
    leakage = _read_csv(processed / "global_walk_forward_leakage_audit.csv")

    random_required = {
        "portfolio_id",
        "benchmark_scope",
        "benchmark_provenance_status",
        "protocol_hash",
        "annualized_return",
        "volatility",
        "sharpe",
        "max_drawdown",
        "cvar_95",
    }
    random_schema_ok = not randoms.empty and random_required.issubset(randoms.columns)
    random_scope_ok = bool(
        random_schema_ok
        and randoms["benchmark_scope"].astype(str).eq("walk_forward_oos_net").all()
        and randoms["benchmark_provenance_status"]
        .astype(str)
        .eq("verified_same_protocol")
        .all()
    )
    model_date_hashes = _date_hashes_by_group(model_returns, "model_name")
    random_date_hashes = _date_hashes_by_group(random_returns, "portfolio_id")
    equal_weight_hash = model_date_hashes.get("Equal Weight", "missing")
    provenance_identity_fields = [
        "run_id",
        "execution_id",
        "data_as_of_date",
        "generated_at",
        "config_hash",
        "input_fingerprint",
        "universe_snapshot_id",
        "data_snapshot_id",
    ]
    provenance_ok = bool(
        random_scope_ok
        and provenance.get("benchmark_scope") == "walk_forward_oos_net"
        and provenance.get("provenance_status") == "verified_same_protocol"
        and provenance.get("oos_dates_match") is True
        and equal_weight_hash != "missing"
        and model_date_hashes
        and random_date_hashes
        and all(value == equal_weight_hash for value in model_date_hashes.values())
        and all(value == equal_weight_hash for value in random_date_hashes.values())
        and provenance.get("model_oos_dates_hash") == equal_weight_hash
        and provenance.get("random_oos_dates_hash") == equal_weight_hash
        and provenance.get("random_weights_hash")
        == _stable_frame_hash(
            random_weights,
            [
                "fold",
                "portfolio_id",
                "ticker",
                "target_weight",
                "pre_trade_weight",
                "post_test_weight",
            ],
        )
        and randoms["protocol_hash"]
        .astype(str)
        .eq(str(provenance.get("protocol_hash")))
        .all()
        and all(
            str(provenance.get(field)) == str(manifest.get(field))
            for field in provenance_identity_fields
        )
    )
    random_sharpe = (
        pd.to_numeric(randoms["sharpe"], errors="coerce")
        if random_schema_ok
        else pd.Series(dtype=float)
    )
    random_non_degenerate = bool(
        random_scope_ok
        and randoms["portfolio_id"].nunique() >= 2
        and random_sharpe.notna().sum() >= 2
        and random_sharpe.nunique(dropna=True) >= 2
    )
    checks.append(
        _check(
            "random_benchmark_is_same_protocol_walk_forward_oos_net",
            provenance_ok,
            (
                f"rows={len(randoms)}; scopes="
                f"{sorted(randoms.get('benchmark_scope', pd.Series(dtype=str)).astype(str).unique())}; "
                f"provenance={provenance.get('provenance_status', 'missing')}; "
                f"date_hash={equal_weight_hash}"
            ),
        )
    )
    checks.append(
        _check(
            "random_benchmark_distribution_is_non_degenerate",
            random_non_degenerate,
            (
                f"portfolios={randoms.get('portfolio_id', pd.Series(dtype=object)).nunique()}; "
                f"unique_sharpe={random_sharpe.nunique(dropna=True)}"
            ),
        )
    )

    scope_column_ok = bool(
        not selection.empty
        and "random_benchmark_scope" in selection
        and "random_benchmark_provenance_status" in selection
        and selection["random_benchmark_scope"]
        .astype(str)
        .eq("walk_forward_oos_net")
        .all()
        and selection["random_benchmark_provenance_status"]
        .astype(str)
        .eq("verified_same_protocol")
        .all()
    )
    checks.append(
        _check(
            "model_selection_uses_walk_forward_oos_random_scope",
            scope_column_ok,
            (
                "scopes="
                f"{sorted(selection.get('random_benchmark_scope', pd.Series(dtype=str)).astype(str).unique())}"
            ),
        )
    )

    metric_pairs = {
        "walk_forward_annualized_return": "oos_annualized_return",
        "walk_forward_volatility": "oos_volatility",
        "walk_forward_sharpe": "oos_sharpe",
        "walk_forward_sortino": "oos_sortino",
        "walk_forward_max_drawdown": "oos_max_drawdown",
        "walk_forward_cvar_95": "oos_cvar_95",
    }
    required_selection = {"model_name", *metric_pairs}
    required_walk = {"model_name", *metric_pairs.values()}
    reconciled = False
    max_difference = float("inf")
    compared_rows = 0
    if (
        not selection.empty
        and not walk.empty
        and required_selection.issubset(selection.columns)
        and required_walk.issubset(walk.columns)
    ):
        merged = selection[list(required_selection)].merge(
            walk[list(required_walk)],
            on="model_name",
            how="inner",
            validate="one_to_one",
        )
        differences: list[float] = []
        for selection_column, walk_column in metric_pairs.items():
            left = pd.to_numeric(merged[selection_column], errors="coerce")
            right = pd.to_numeric(merged[walk_column], errors="coerce")
            finite = left.notna() & right.notna()
            differences.extend((left.loc[finite] - right.loc[finite]).abs().tolist())
        compared_rows = len(merged)
        max_difference = max(differences, default=float("inf"))
        reconciled = bool(compared_rows > 0 and max_difference <= 1e-10)
    checks.append(
        _check(
            "model_selection_metrics_reconcile_to_walk_forward_oos",
            reconciled,
            f"models_compared={compared_rows}; max_abs_difference={max_difference}",
        )
    )
    leakage_required = {
        "fold",
        "check",
        "passed",
        "audit_status",
        "evidence_scope",
        *provenance_identity_fields,
    }
    leakage_checks = {
        "train_end_before_test_start",
        "scores_as_of_not_after_train_end",
        "selected_tickers_available_in_train",
        "scores_recomputed_inside_fold",
    }
    leakage_identity_ok = bool(
        not leakage.empty
        and leakage_required.issubset(leakage.columns)
        and all(
            set(leakage[field].dropna().astype(str)) == {str(manifest.get(field))}
            for field in provenance_identity_fields
        )
    )
    leakage_fold_sets_ok = bool(
        leakage_identity_ok
        and not leakage.duplicated(["fold", "check"]).any()
        and all(
            set(group["check"].astype(str)) == leakage_checks
            for _, group in leakage.groupby("fold", sort=False)
        )
    )
    leakage_passed = bool(
        leakage_fold_sets_ok
        and leakage["passed"].map(_truthy).all()
        and leakage["audit_status"]
        .astype(str)
        .eq("passed_with_current_universe_survivorship_limitation")
        .all()
        and leakage["evidence_scope"]
        .astype(str)
        .eq("current_universe_not_point_in_time")
        .all()
    )
    selection_leakage_passed = bool(
        not selection.empty
        and {
            "leakage_gate_pass",
            "leakage_evidence_status",
        }.issubset(selection.columns)
        and selection["leakage_gate_pass"].map(_truthy).all()
        and selection["leakage_evidence_status"]
        .astype(str)
        .eq("verified_current_no_lookahead_with_survivorship_limitation")
        .all()
    )
    checks.append(
        _check(
            "walk_forward_leakage_gate_is_current_complete_and_fail_closed",
            leakage_passed and selection_leakage_passed,
            (
                f"audit_rows={len(leakage)}; identity_ok={leakage_identity_ok}; "
                f"fold_sets_ok={leakage_fold_sets_ok}; "
                f"selection_gate={selection_leakage_passed}"
            ),
        )
    )
    uncertainty_required = {
        "model_name",
        "uncertainty_status",
        "uncertainty_method",
        "paired_observations",
        "sharpe_diff_ci_lower",
        "sharpe_diff_ci_upper",
        "probability_sharpe_improvement",
    }
    uncertainty_schema_ok = bool(
        not uncertainty.empty
        and uncertainty_required.issubset(uncertainty.columns)
        and uncertainty["uncertainty_method"]
        .astype(str)
        .eq("paired_circular_block_bootstrap")
        .all()
    )
    checks.append(
        _check(
            "walk_forward_paired_block_uncertainty_present",
            uncertainty_schema_ok,
            (
                f"rows={len(uncertainty)}; "
                f"missing={sorted(uncertainty_required.difference(uncertainty.columns))}"
            ),
        )
    )
    robustness_is_promotion_grade = bool(
        robustness.get("robustness_status") == "promotion_grade_nested_walk_forward_oos"
        and robustness.get("robustness_method")
        == "nested_chronological_walk_forward_oos"
        and _truthy(robustness.get("promotion_eligible"))
        and all(
            str(robustness.get(field)) == str(manifest.get(field))
            for field in provenance_identity_fields
        )
    )
    reported_robustness_gate = (
        selection["robustness_gate_pass"].map(_truthy)
        if "robustness_gate_pass" in selection
        else pd.Series([True])
    )
    checks.append(
        _check(
            "robustness_promotion_gate_fails_closed",
            bool(robustness_is_promotion_grade or not reported_robustness_gate.any()),
            (
                f"robustness_status={robustness.get('robustness_status', 'missing')}; "
                f"promotion_grade={robustness_is_promotion_grade}; "
                f"reported_passes={int(reported_robustness_gate.sum())}"
            ),
        )
    )


def _check_publication_source_registry(
    root: Path,
    processed: Path,
    checks: list[dict[str, object]],
) -> None:
    manifest = _read_json(processed / "quantverse_v2_run_manifest.json")
    failures = validate_registered_artifacts(
        processed,
        [
            *[processed / name for name in PUBLICATION_SOURCE_ARTIFACTS],
            *[root / name for name in ADDITIONAL_PUBLICATION_SOURCE_ARTIFACTS],
        ],
        manifest=manifest,
        root=root,
    )
    checks.append(
        _check(
            "publication_sources_match_registered_sha256_and_run_identity",
            not failures,
            f"failures={failures}",
        )
    )


def _check_demo_run_status(
    summary: dict[str, object], checks: list[dict[str, object]]
) -> None:
    status = str(summary.get("run_status", "missing")).strip().lower()
    checks.append(
        _check(
            "demo_run_completed_without_failed_step",
            status == "completed" and not str(summary.get("failed_step", "")).strip(),
            (
                f"run_status={status}; "
                f"failed_step={summary.get('failed_step', 'none')}"
            ),
        )
    )


def _check_final_model_consistency(
    summary: dict[str, object],
    decision: dict[str, object],
    league: pd.DataFrame,
    selection: pd.DataFrame,
    checks: list[dict[str, object]],
) -> None:
    summary_model = str(summary.get("final_selected_model", "")).strip()
    decision_model = str(decision.get("final_selected_model", "")).strip()
    league_models = (
        set(league["model_name"].astype(str)) if "model_name" in league else set()
    )
    checks.append(
        _check(
            "summary_matches_final_model_decision",
            bool(summary_model and summary_model == decision_model),
            f"summary={summary_model}; decision={decision_model}",
        )
    )
    checks.append(
        _check(
            "final_model_appears_in_model_league",
            bool(summary_model and summary_model in league_models),
            f"final_model={summary_model}",
        )
    )
    required = {
        "model_name",
        "eligible_final_model",
        "selection_score",
        "book_grounded_rank",
        "random_sharpe_percentile",
        "promotion_gate_failed_reasons",
        "sharpe_improvement_vs_equal_weight",
        "beats_equal_weight_sharpe",
        "drawdown_not_materially_worse_than_equal_weight",
        "cvar_not_materially_worse_than_equal_weight",
        "turnover_within_limit",
        "random_sharpe_gate_pass",
        "uncertainty_gate_pass",
        "robustness_gate_pass",
        "forecast_validation_gate_pass",
        "extreme_metric_warning",
        "walk_forward_sharpe",
        "walk_forward_annualized_return",
        "walk_forward_max_drawdown",
        "walk_forward_cvar_95",
        "turnover",
    }
    schema_ok = bool(not selection.empty and required.issubset(selection.columns))
    final_rows = (
        selection.loc[selection["model_name"].astype(str).eq(decision_model)]
        if schema_ok
        else pd.DataFrame()
    )
    unique_eligible = bool(
        len(final_rows) == 1 and _truthy(final_rows["eligible_final_model"].iloc[0])
    )
    checks.append(
        _check(
            "final_model_is_unique_eligible_selection_row",
            unique_eligible,
            (
                f"schema_ok={schema_ok}; final_model={decision_model}; "
                f"matching_rows={len(final_rows)}"
            ),
        )
    )

    expected_row = _independent_expected_final_selection_row(selection)
    decision_reconciles = bool(
        unique_eligible
        and expected_row is not None
        and decision_model == str(expected_row["model_name"])
        and str(decision.get("final_model_selection_method", ""))
        == "paired_block_bootstrap_gate_then_oos_sharpe"
        and _optional_number_matches(
            decision.get("final_model_selection_score"),
            expected_row.get("selection_score"),
        )
        and _optional_number_matches(
            decision.get("random_portfolio_percentile"),
            expected_row.get("random_sharpe_percentile"),
        )
        and _optional_number_matches(
            decision.get("final_model_book_grounded_rank"),
            expected_row.get("book_grounded_rank"),
        )
        and str(decision.get("final_model_gate_reasons", ""))
        == str(expected_row.get("promotion_gate_failed_reasons", ""))
    )
    checks.append(
        _check(
            "final_model_decision_reconciles_to_selection_evidence",
            decision_reconciles,
            (
                f"decision_model={decision_model}; expected_model="
                f"{expected_row.get('model_name') if expected_row is not None else 'missing'}"
            ),
        )
    )
    checks.append(
        _check(
            "final_decision_is_fail_closed_for_public_data_scope",
            str(decision.get("final_decision", "")).strip().lower() == "not promoted",
            f"final_decision={decision.get('final_decision', 'missing')}",
        )
    )


def _independent_expected_final_selection_row(
    selection: pd.DataFrame,
) -> pd.Series | None:
    required = {
        "model_name",
        "eligible_final_model",
        "sharpe_improvement_vs_equal_weight",
        "beats_equal_weight_sharpe",
        "drawdown_not_materially_worse_than_equal_weight",
        "cvar_not_materially_worse_than_equal_weight",
        "turnover_within_limit",
        "random_sharpe_gate_pass",
        "uncertainty_gate_pass",
        "robustness_gate_pass",
        "forecast_validation_gate_pass",
        "extreme_metric_warning",
        "walk_forward_sharpe",
        "walk_forward_annualized_return",
        "walk_forward_max_drawdown",
        "walk_forward_cvar_95",
        "turnover",
    }
    if selection.empty or not required.issubset(selection.columns):
        return None
    eligible = selection.loc[selection["eligible_final_model"].map(_truthy)].copy()
    equal_weight = eligible.loc[eligible["model_name"].astype(str).eq("Equal Weight")]
    if len(equal_weight) != 1:
        return None
    active = eligible.loc[~eligible["model_name"].astype(str).eq("Equal Weight")].copy()
    if active.empty:
        return equal_weight.iloc[0]
    active_gate = (
        pd.to_numeric(active["sharpe_improvement_vs_equal_weight"], errors="coerce").ge(
            0.0
        )
        & active["beats_equal_weight_sharpe"].map(_truthy)
        & active["drawdown_not_materially_worse_than_equal_weight"].map(_truthy)
        & active["cvar_not_materially_worse_than_equal_weight"].map(_truthy)
        & active["turnover_within_limit"].map(_truthy)
        & active["random_sharpe_gate_pass"].map(_truthy)
        & active["uncertainty_gate_pass"].map(_truthy)
        & active["robustness_gate_pass"].map(_truthy)
        & active["forecast_validation_gate_pass"].map(_truthy)
        & active["extreme_metric_warning"]
        .astype(str)
        .str.strip()
        .str.lower()
        .eq("none")
    )
    if not active_gate.any():
        return equal_weight.iloc[0]
    return (
        active.loc[active_gate]
        .sort_values(
            [
                "walk_forward_sharpe",
                "walk_forward_annualized_return",
                "walk_forward_max_drawdown",
                "walk_forward_cvar_95",
                "turnover",
                "model_name",
            ],
            ascending=[False, False, False, False, True, True],
        )
        .iloc[0]
    )


def _optional_number_matches(observed: object, expected: object) -> bool:
    try:
        observed_missing = observed is None or bool(pd.isna(observed))
        expected_missing = expected is None or bool(pd.isna(expected))
    except (TypeError, ValueError):
        return False
    if observed_missing or expected_missing:
        return observed_missing and expected_missing
    try:
        return bool(
            np.isclose(
                float(observed),
                float(expected),
                atol=1e-12,
                rtol=1e-12,
            )
        )
    except (TypeError, ValueError):
        return False


def _check_final_weights(
    summary: dict[str, object],
    weights: pd.DataFrame,
    checks: list[dict[str, object]],
) -> None:
    model = str(summary.get("final_selected_model", "")).strip()
    selected = (
        weights.loc[weights["model_name"].astype(str).eq(model)]
        if not weights.empty and "model_name" in weights
        else pd.DataFrame()
    )
    weight_sum = (
        float(pd.to_numeric(selected["weight"], errors="coerce").sum())
        if "weight" in selected
        else 0.0
    )
    selected_count = int(
        (
            pd.to_numeric(
                selected.get("weight", pd.Series(dtype=float)), errors="coerce"
            ).abs()
            > 1e-8
        ).sum()
    )
    checks.append(
        _check(
            "final_weights_sum_to_one",
            abs(weight_sum - 1.0) <= 1e-6,
            f"model={model}; weight_sum={weight_sum}",
        )
    )
    checks.append(
        _check(
            "selected_holdings_count_plausible",
            1 <= selected_count <= 100,
            f"selected_count={selected_count}",
        )
    )


def _check_pdf(
    path: Path, checks: list[dict[str, object]], label: str = "report_pdf"
) -> None:
    try:
        reader = PdfReader(str(path))
        pages = len(reader.pages)
        first_text = reader.pages[0].extract_text() if pages else ""
        passed = path.exists() and pages > 0 and bool(first_text)
        details = f"pages={pages}; first_text_chars={len(first_text or '')}"
    except Exception as exc:
        passed = False
        details = _portable_exception_details(exc, path)
    checks.append(_check(label, passed, details))


def _check_html(path: Path, checks: list[dict[str, object]]) -> None:
    text = path.read_text(encoding="utf-8", errors="ignore") if path.exists() else ""
    missing = [section for section in REQUIRED_HTML_SECTIONS if section not in text]
    checks.append(
        _check(
            "html_required_sections",
            path.exists() and not missing,
            f"missing_sections={missing}",
        )
    )


def _check_excel(path: Path, checks: list[dict[str, object]]) -> None:
    try:
        sheets = _excel_sheet_names(path)
        missing = [sheet for sheet in REQUIRED_EXCEL_SHEETS if sheet not in sheets]
        passed = path.exists() and not missing
        details = f"missing_sheets={missing}; sheet_count={len(sheets)}"
    except Exception as exc:
        passed = False
        details = _portable_exception_details(exc, path)
    checks.append(_check("excel_required_sheets", passed, details))


def _check_excel_formula_errors(
    path: Path,
    checks: list[dict[str, object]],
) -> None:
    error_cells: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                if not name.startswith("xl/worksheets/sheet") or not name.endswith(
                    ".xml"
                ):
                    continue
                root = ElementTree.fromstring(archive.read(name))
                namespace = {
                    "main": (
                        "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                    )
                }
                for cell in root.findall(".//main:c[@t='e']", namespace):
                    value = cell.find("main:v", namespace)
                    error_cells.append(
                        f"{name}:{cell.attrib.get('r')}="
                        f"{value.text if value is not None else 'unknown'}"
                    )
        passed = path.exists() and not error_cells
        details = f"formula_error_count={len(error_cells)}; errors={error_cells[:10]}"
    except Exception as exc:
        passed = False
        details = _portable_exception_details(exc, path)
    checks.append(_check("excel_no_formula_errors", passed, details))


def _check_publication_manifests(
    root: Path,
    summary: dict[str, object],
    checks: list[dict[str, object]],
) -> None:
    expected_run_id = str(summary.get("run_id", "")).strip()
    manifests = [
        (
            "report",
            root / "output" / "quantverse_v2_report_publication_manifest.json",
            REPORT_PUBLICATION_TYPE,
            REPORT_PUBLICATION_ARTIFACTS,
        ),
        (
            "excel",
            root / "output" / "quantverse_v2_excel_publication_manifest.json",
            EXCEL_PUBLICATION_TYPE,
            EXCEL_PUBLICATION_ARTIFACTS,
        ),
    ]
    expected_identity = {field: summary.get(field, "") for field in RUN_FIELDS}
    for prefix, path, publication_type, expected_artifacts in manifests:
        manifest_checks = validate_publication_manifest(
            root,
            path,
            expected_run_id=expected_run_id,
            expected_publication_type=publication_type,
            expected_artifacts=expected_artifacts,
            expected_run_identity=expected_identity,
        )
        for check in manifest_checks:
            checks.append(
                _check(
                    f"{prefix}_{check['check']}",
                    bool(check["passed"]),
                    str(check["details"]),
                )
            )


def _check_report_claim_language(root: Path, checks: list[dict[str, object]]) -> None:
    paths = [
        root / "output" / "pdf" / "quantverse_v2_research_report.pdf",
        root / "output" / "pdf" / "quantverse_v2_executive_research_report.pdf",
        root / "output" / "pdf" / "quantverse_v2_methodology_validation_appendix.pdf",
        root / "output" / "html" / "quantverse_v2_research_report.html",
    ]
    text = "\n".join(_extract_text(path) for path in paths)
    hits = []
    for pattern in FORBIDDEN_REPORT_PATTERNS:
        if re.search(pattern, text, flags=re.IGNORECASE):
            hits.append(pattern)
    checks.append(
        _check(
            "generated_reports_no_forbidden_claims",
            not hits,
            f"forbidden_hits={hits}",
        )
    )


def _check_numerical_integrity(
    root: Path,
    summary: dict[str, object],
    checks: list[dict[str, object]],
) -> None:
    result = validate_v2_numerical_integrity(root)
    for check in result["checks"]:
        checks.append(
            _check(
                f"numerical_integrity_{check['check']}",
                bool(check["passed"]),
                str(check["details"]),
            )
        )
    summary_status = str(summary.get("numerical_integrity_status", "")).strip()
    summary_failed = summary.get("numerical_integrity_failed_checks")
    try:
        summary_failed_count = int(summary_failed)
    except (TypeError, ValueError):
        summary_failed_count = -1
    checks.append(
        _check(
            "summary_numerical_integrity_matches_artifact_validation",
            summary_status == result["overall_status"]
            and summary_failed_count == int(result["failed_check_count"]),
            (
                f"summary_status={summary_status}; actual_status={result['overall_status']}; "
                f"summary_failed={summary_failed}; actual_failed={result['failed_check_count']}"
            ),
        )
    )


def _check_reference_math(processed: Path, checks: list[dict[str, object]]) -> None:
    summary = _read_json(processed / "quantverse_v2_reference_math_summary.json")
    frame = _read_csv(processed / "quantverse_v2_reference_math_checks.csv")
    manifest = _read_json(processed / "quantverse_v2_run_manifest.json")
    if not summary or frame.empty:
        checks.append(
            _check(
                "independent_reference_math_passed",
                False,
                "Independent reference-math summary or checks are missing.",
            )
        )
        return
    passed_values = frame.get("passed", pd.Series(dtype=bool)).map(
        lambda value: str(value).strip().lower() in {"1", "true", "yes"}
    )
    failed = int((~passed_values).sum())
    run_ids = (
        frame.get("run_id", pd.Series(dtype=str)).dropna().astype(str).unique().tolist()
    )
    manifest_run_id = str(manifest.get("run_id", "missing"))
    summary_run_id = str(summary.get("run_id", "missing"))
    passed = bool(
        str(summary.get("status")) == "passed"
        and int(summary.get("failed_check_count", -1)) == 0
        and int(summary.get("check_count", -1)) == len(frame)
        and failed == 0
        and run_ids == [manifest_run_id]
        and summary_run_id == manifest_run_id
    )
    checks.append(
        _check(
            "independent_reference_math_passed",
            passed,
            (
                f"summary_status={summary.get('status')}; checks={len(frame)}; "
                f"failed={failed}; run_ids={run_ids}; "
                f"manifest_run_id={manifest_run_id}"
            ),
        )
    )


def _check_visual_analytics(root: Path, checks: list[dict[str, object]]) -> None:
    result = validate_visual_analytics_outputs(root / "data" / "processed")
    for check in result["checks"]:
        checks.append(
            _check(
                f"visual_analytics_{check['check']}",
                bool(check["passed"]),
                str(check["details"]),
            )
        )


def _check_exposure_metadata_quality(
    processed: Path, checks: list[dict[str, object]]
) -> None:
    quality = _read_csv(processed / "global_exposure_metadata_quality.csv")
    if quality.empty:
        checks.append(
            _check(
                "exposure_metadata_quality_report_present",
                False,
                "missing global_exposure_metadata_quality.csv",
            )
        )
        return
    required = {
        "exposure_metadata_status",
        "sector_coverage_ratio",
        "industry_coverage_ratio",
        "issuer_country_coverage_ratio",
        "economic_country_coverage_ratio",
        "listing_country_coverage_ratio",
        "metadata_confidence_distribution",
        "listing_country_vs_issuer_country_warning",
    }
    missing = sorted(required.difference(quality.columns))
    status = str(quality["exposure_metadata_status"].iloc[0])
    sector = _float(quality["sector_coverage_ratio"].iloc[0])
    industry = _float(quality["industry_coverage_ratio"].iloc[0])
    issuer = _float(quality["issuer_country_coverage_ratio"].iloc[0])
    economic = _float(quality["economic_country_coverage_ratio"].iloc[0])
    listing = _float(quality["listing_country_coverage_ratio"].iloc[0])
    valid_statuses = {
        "passed",
        "passed_with_metadata_warning",
        "diagnostic_metadata_incomplete",
        "failed",
    }
    checks.append(
        _check(
            "exposure_metadata_quality_schema",
            not missing,
            f"missing_columns={missing}",
        )
    )
    checks.append(
        _check(
            "exposure_metadata_incomplete_is_not_plain_pass",
            bool(
                status in valid_statuses
                and not (
                    (sector == 0.0 or issuer == 0.0)
                    and status in {"passed", "passed_with_metadata_warning"}
                )
            ),
            (
                f"status={status}; sector_coverage_ratio={sector}; "
                f"industry_coverage_ratio={industry}; "
                f"issuer_country_coverage_ratio={issuer}; "
                f"economic_country_coverage_ratio={economic}; "
                f"listing_country_coverage_ratio={listing}"
            ),
        )
    )
    checks.append(
        _check(
            "listing_issuer_economic_exposure_files_exist",
            all(
                (processed / filename).exists()
                for filename in [
                    "global_listing_country_exposure.csv",
                    "global_issuer_country_exposure.csv",
                    "global_economic_country_exposure.csv",
                    "global_industry_exposure.csv",
                    "global_exchange_exposure.csv",
                ]
            ),
            "separate listing/issuer/economic/industry/exchange exposure files are required",
        )
    )
    holdings = _read_csv(processed / "global_top_holdings_explanation.csv")
    if not holdings.empty and "adr_or_foreign_issuer_flag" in holdings:
        flagged = holdings.loc[
            holdings["adr_or_foreign_issuer_flag"].map(
                lambda value: str(value).lower() == "true"
            )
        ]
        bad = flagged.loc[
            flagged.get("issuer_country", pd.Series(index=flagged.index, dtype=object))
            .fillna("")
            .astype(str)
            .str.strip()
            .eq(
                flagged.get(
                    "listing_country", pd.Series(index=flagged.index, dtype=object)
                )
                .fillna("")
                .astype(str)
                .str.strip()
            )
        ]
        checks.append(
            _check(
                "foreign_issuer_not_collapsed_to_listing_country",
                bad.empty,
                f"flagged_holdings={len(flagged)}; collapsed_rows={len(bad)}",
            )
        )


def _check_security_identity_history(
    processed: Path,
    checks: list[dict[str, object]],
) -> None:
    identity = _read_csv(processed / "global_security_identity_audit.csv")
    history = _read_csv(processed / "global_security_history_eligibility.csv")
    features = _read_csv(processed / "global_feature_history_eligibility.csv")
    scores = _read_csv(processed / "global_stock_scores.csv")
    reconciliation = _read_csv(
        processed / "global_cross_artifact_count_reconciliation.csv"
    )
    registry = _read_csv(processed / "quantverse_v2_artifact_run_registry.csv")
    manifest = _read_json(processed / "quantverse_v2_run_manifest.json")

    identity_required = {
        "ticker",
        "current_listing_start_date",
        "provider_history_start_date",
        "first_valid_return_date",
        "observations_before_current_listing",
        "ticker_reuse_status",
        "identity_continuity_status",
        "history_contamination_status",
        "standard_scoring_eligible",
        "forecast_eligible",
        "walk_forward_eligible",
        "run_id",
    }
    checks.append(
        _check(
            "security_identity_audit_present",
            not identity.empty and identity_required.issubset(identity.columns),
            f"rows={len(identity)}; missing={sorted(identity_required.difference(identity.columns))}",
        )
    )
    feature_required = {
        "ticker",
        "observations",
        "12m_eligible",
        "volatility_12m_eligible",
        "standard_composite_score_eligible",
        "eligibility_status",
        "run_id",
    }
    checks.append(
        _check(
            "feature_history_eligibility_audit_present",
            not features.empty and feature_required.issubset(features.columns),
            f"rows={len(features)}; missing={sorted(feature_required.difference(features.columns))}",
        )
    )
    history_required = {
        "ticker",
        "eligibility_status",
        "standard_scoring_eligible",
        "forecast_eligible",
        "walk_forward_eligible",
        "run_id",
    }
    checks.append(
        _check(
            "security_history_eligibility_audit_present",
            not history.empty and history_required.issubset(history.columns),
            f"rows={len(history)}; missing={sorted(history_required.difference(history.columns))}",
        )
    )

    if identity.empty or not {
        "ticker",
        "ticker_reuse_status",
        "identity_continuity_status",
        "observations_before_current_listing",
        "history_contamination_status",
    }.issubset(identity):
        unresolved_reuse = pd.DataFrame()
        pre_listing_failures = pd.DataFrame()
    else:
        reuse = (
            identity["ticker_reuse_status"]
            .astype(str)
            .str.contains("known_reuse", case=False, na=False)
        )
        continuity_verified = (
            identity["identity_continuity_status"]
            .astype(str)
            .str.startswith("verified")
        )
        unresolved_reuse = identity.loc[reuse & ~continuity_verified]
        pre_listing = pd.to_numeric(
            identity["observations_before_current_listing"], errors="coerce"
        )
        contamination_resolved = (
            identity["history_contamination_status"]
            .astype(str)
            .isin(
                [
                    "none_detected",
                    "detected_and_removed",
                    "verified_continuity_preserved",
                ]
            )
        )
        pre_listing_failures = identity.loc[(pre_listing > 0) & ~contamination_resolved]
        listing_dates = pd.to_datetime(
            identity["current_listing_start_date"],
            format="%Y-%m-%d",
            errors="coerce",
        )
        first_return_dates = pd.to_datetime(
            identity["first_valid_return_date"],
            format="%Y-%m-%d",
            errors="coerce",
        )
        preserved_continuity = (
            identity["identity_continuity_status"]
            .astype(str)
            .isin(
                [
                    "verified_same_security_continuity",
                    "verified_predecessor_continuity",
                ]
            )
        )
        invalid_return_start = (
            listing_dates.notna()
            & first_return_dates.notna()
            & first_return_dates.lt(listing_dates)
            & ~preserved_continuity
        )
        pre_listing_failures = pd.concat(
            [pre_listing_failures, identity.loc[invalid_return_start]],
            ignore_index=True,
        ).drop_duplicates("ticker")
    checks.append(
        _check(
            "no_unresolved_ticker_reuse_in_standard_selection",
            unresolved_reuse.empty,
            f"unresolved_tickers={unresolved_reuse.get('ticker', pd.Series(dtype=str)).astype(str).tolist()}",
        )
    )
    checks.append(
        _check(
            "no_pre_listing_history_contamination",
            pre_listing_failures.empty,
            f"failed_tickers={pre_listing_failures.get('ticker', pd.Series(dtype=str)).astype(str).tolist()}",
        )
    )

    score_feature_columns = {
        "ticker",
        "selection_flag",
        "standard_composite_score_eligible",
    }
    feature_contract_columns = {
        "ticker",
        "observations",
        "12m_eligible",
        "volatility_12m_eligible",
        "standard_composite_score_eligible",
    }
    feature_contract_available = bool(
        not scores.empty
        and not features.empty
        and score_feature_columns.issubset(scores)
        and feature_contract_columns.issubset(features)
    )
    if not feature_contract_available:
        selected_short = pd.DataFrame()
        feature_failures = pd.DataFrame()
    else:
        score_flags = scores[
            ["ticker", "selection_flag", "standard_composite_score_eligible"]
        ].copy()
        merged = score_flags.merge(
            features[
                [
                    "ticker",
                    "observations",
                    "12m_eligible",
                    "volatility_12m_eligible",
                    "standard_composite_score_eligible",
                ]
            ],
            on="ticker",
            how="left",
            suffixes=("_score", "_feature"),
        )
        selected_mask = merged["selection_flag"].map(_truthy)
        standard_mask = merged["standard_composite_score_eligible_score"].map(_truthy)
        selected_short = merged.loc[selected_mask & ~standard_mask]
        feature_failures = merged.loc[
            standard_mask
            & (
                ~merged["12m_eligible"].map(_truthy)
                | ~merged["volatility_12m_eligible"].map(_truthy)
                | ~merged["standard_composite_score_eligible_feature"].map(_truthy)
                | pd.to_numeric(merged["observations"], errors="coerce")
                .lt(252)
                .fillna(True)
            )
        ]
    checks.append(
        _check(
            "feature_history_sufficiency_valid",
            feature_contract_available and feature_failures.empty,
            (
                "required score/feature columns missing"
                if not feature_contract_available
                else f"failed_tickers={feature_failures.get('ticker', pd.Series(dtype=str)).astype(str).tolist()}"
            ),
        )
    )
    checks.append(
        _check(
            "short_history_assets_not_silently_promoted",
            feature_contract_available and selected_short.empty,
            (
                "required score/feature columns missing"
                if not feature_contract_available
                else f"selected_short_history={selected_short.get('ticker', pd.Series(dtype=str)).astype(str).tolist()}"
            ),
        )
    )
    ineligible_tickers = _ineligible_feature_tickers(features)
    portfolio_input_violations = _portfolio_input_violations(
        processed, ineligible_tickers
    )
    checks.append(
        _check(
            "no_short_history_assets_in_portfolio_inputs",
            feature_contract_available and not portfolio_input_violations,
            f"violations={portfolio_input_violations}",
        )
    )
    reuse_tickers = (
        set(
            identity.loc[
                identity.get("ticker_reuse_status", pd.Series(dtype=str))
                .astype(str)
                .str.contains("known_reuse", case=False, na=False),
                "ticker",
            ].astype(str)
        )
        if not identity.empty and "ticker" in identity
        else set()
    )
    reuse_violations = sorted(reuse_tickers.intersection(portfolio_input_violations))
    checks.append(
        _check(
            "no_ticker_reuse_warning_ignored_in_portfolio_inputs",
            not reuse_violations,
            f"violations={reuse_violations}",
        )
    )

    reconciliation_pass = bool(
        not reconciliation.empty
        and "status" in reconciliation
        and reconciliation["status"].astype(str).eq("passed").all()
    )
    checks.append(
        _check(
            "selected_forecast_count_reconciled",
            reconciliation_pass
            and _reconciliation_row_passed(
                reconciliation, "forecast_output_ticker_count"
            ),
            _reconciliation_details(reconciliation, "forecast_output_ticker_count"),
        )
    )
    checks.append(
        _check(
            "final_holdings_count_reconciled",
            reconciliation_pass
            and _reconciliation_row_passed(reconciliation, "final_model_holding_count"),
            _reconciliation_details(reconciliation, "final_model_holding_count"),
        )
    )
    checks.append(
        _check(
            "cross_artifact_count_reconciliation_passed",
            reconciliation_pass,
            f"failed={_failed_reconciliation_artifacts(reconciliation)}",
        )
    )

    expected_run_id = str(manifest.get("run_id", "")).strip()
    core_registry = (
        registry.loc[
            registry.get("artifact", pd.Series(dtype=str))
            .astype(str)
            .isin(
                [
                    "data/processed/global_security_identity_audit.csv",
                    "data/processed/global_feature_history_eligibility.csv",
                    "data/processed/global_stock_scores.csv",
                    "data/processed/global_stock_return_forecasts.csv",
                    "data/processed/global_portfolio_league_weights.csv",
                    "data/processed/global_portfolio_risk_report.csv",
                    "data/processed/global_final_model_decision.json",
                    "data/processed/global_robustness_sensitivity.csv",
                    "data/processed/global_exposure_metadata_quality.csv",
                    "data/processed/global_walk_forward_window_summary.csv",
                ]
            )
        ]
        if not registry.empty and {"artifact", "run_id"}.issubset(registry)
        else pd.DataFrame()
    )
    registered_ids = (
        set(core_registry["run_id"].dropna().astype(str))
        if "run_id" in core_registry
        else set()
    )
    run_ids_consistent = bool(
        expected_run_id
        and len(core_registry) == 10
        and registered_ids == {expected_run_id}
    )
    checks.append(
        _check(
            "generated_artifact_run_ids_consistent",
            run_ids_consistent,
            f"expected={expected_run_id}; registered={sorted(registered_ids)}; core_rows={len(core_registry)}",
        )
    )


def _ineligible_feature_tickers(features: pd.DataFrame) -> set[str]:
    if features.empty or not {
        "ticker",
        "standard_composite_score_eligible",
    }.issubset(features):
        return set()
    eligible = features["standard_composite_score_eligible"].map(_truthy)
    return set(features.loc[~eligible, "ticker"].astype(str))


def _portfolio_input_violations(
    processed: Path,
    ineligible_tickers: set[str],
) -> list[str]:
    if not ineligible_tickers:
        return []
    violations: set[str] = set()
    for filename in [
        "global_portfolio_league_weights.csv",
        "global_master_candidate_weights.csv",
    ]:
        frame = _read_csv(processed / filename)
        columns = {str(column).lower(): column for column in frame.columns}
        ticker_column = columns.get("ticker")
        weight_column = columns.get("weight")
        if ticker_column is None or weight_column is None:
            continue
        positive = pd.to_numeric(frame[weight_column], errors="coerce").abs().gt(1e-12)
        violations.update(
            set(frame.loc[positive, ticker_column].astype(str)).intersection(
                ineligible_tickers
            )
        )

    selected = _read_csv(processed / "global_master_selected_assets.csv")
    selected_columns = {str(column).lower(): column for column in selected.columns}
    selected_ticker = selected_columns.get("ticker")
    if selected_ticker is not None:
        violations.update(
            set(selected[selected_ticker].astype(str)).intersection(ineligible_tickers)
        )

    contributions = _read_csv(processed / "global_risk_contribution_report.csv")
    contribution_columns = {
        str(column).lower(): column for column in contributions.columns
    }
    contribution_ticker = contribution_columns.get("ticker")
    if contribution_ticker is not None:
        violations.update(
            set(contributions[contribution_ticker].astype(str)).intersection(
                ineligible_tickers
            )
        )

    randoms = _read_csv(processed / "global_master_random_portfolio_benchmark.csv")
    for column in randoms.columns:
        label = str(column)
        if not label.lower().startswith("weight_"):
            continue
        ticker = label[len("weight_") :]
        if ticker in ineligible_tickers:
            violations.add(ticker)

    correlation = _read_csv(processed / "global_correlation_matrix.csv")
    correlation_tickers = {str(column) for column in correlation.columns}
    if not correlation.empty:
        first = correlation.columns[0]
        correlation_tickers.update(correlation[first].dropna().astype(str))
    violations.update(correlation_tickers.intersection(ineligible_tickers))
    return sorted(violations)


def _reconciliation_row_passed(frame: pd.DataFrame, artifact: str) -> bool:
    if frame.empty or not {"artifact", "status"}.issubset(frame):
        return False
    row = frame.loc[frame["artifact"].astype(str).eq(artifact)]
    return bool(not row.empty and row["status"].astype(str).eq("passed").all())


def _reconciliation_details(frame: pd.DataFrame, artifact: str) -> str:
    if frame.empty or "artifact" not in frame:
        return "reconciliation missing"
    row = frame.loc[frame["artifact"].astype(str).eq(artifact)]
    if row.empty:
        return f"{artifact} row missing"
    return str(row.iloc[0].get("observed_relationship", ""))


def _failed_reconciliation_artifacts(frame: pd.DataFrame) -> list[str]:
    if frame.empty or not {"artifact", "status"}.issubset(frame):
        return ["missing"]
    return (
        frame.loc[frame["status"].astype(str).eq("failed"), "artifact"]
        .astype(str)
        .tolist()
    )


def _truthy(value: object) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def validate_selected_stock_report_semantics(root: Path) -> dict[str, object]:
    """Validate the selected-stock table in its exact report-facing contexts."""

    checks: list[dict[str, object]] = []
    processed = root / "data" / "processed"
    scores = _read_csv(processed / "global_stock_scores.csv")
    view = _read_csv(processed / "global_selected_stocks_report_view.csv")
    quality = _read_csv(processed / "global_selected_stocks_report_view_quality.csv")
    exposure_quality = _read_csv(processed / "global_exposure_metadata_quality.csv")

    required_columns = {
        "ticker",
        "selection_rank",
        "composite_quant_score",
        "listing_country",
        "issuer_country",
        "economic_country",
        "listing_currency",
        "exchange",
        "sector",
        "industry",
        "metadata_source",
        "metadata_confidence",
        "adr_or_foreign_issuer_flag",
        "warning_flags",
        "selection_reason",
    }
    missing_columns = sorted(required_columns.difference(view.columns))
    checks.append(
        _check(
            "selected_stock_report_view_non_empty",
            not view.empty and not missing_columns,
            f"rows={len(view)}; missing_columns={missing_columns}",
        )
    )

    selected_count = _selected_score_count(scores)
    checks.append(
        _check(
            "selected_stock_report_view_row_count_matches",
            len(view) == selected_count,
            f"report_view_rows={len(view)}; selected_score_rows={selected_count}",
        )
    )
    duplicate_count = (
        int(view["ticker"].astype(str).str.strip().str.upper().duplicated().sum())
        if "ticker" in view
        else len(view)
    )
    checks.append(
        _check(
            "selected_stock_report_view_no_duplicate_tickers",
            duplicate_count == 0,
            f"duplicate_ticker_count={duplicate_count}",
        )
    )

    quality_required = {
        "selected_stock_count",
        "matched_metadata_count",
        "unmatched_metadata_count",
        "duplicate_ticker_count",
        "semantic_view_status",
    }
    quality_missing = sorted(quality_required.difference(quality.columns))
    if quality.empty or quality_missing:
        quality_passed = False
        quality_details = f"rows={len(quality)}; missing_columns={quality_missing}"
    else:
        row = quality.iloc[0]
        quality_selected = int(_float(row["selected_stock_count"]))
        matched = int(_float(row["matched_metadata_count"]))
        unmatched = int(_float(row["unmatched_metadata_count"]))
        quality_duplicates = int(_float(row["duplicate_ticker_count"]))
        quality_passed = bool(
            quality_selected == selected_count
            and matched + unmatched == selected_count
            and quality_duplicates == 0
        )
        quality_details = (
            f"selected={quality_selected}; matched={matched}; unmatched={unmatched}; "
            f"duplicates={quality_duplicates}; status={row['semantic_view_status']}"
        )
    checks.append(
        _check(
            "selected_stock_report_view_metadata_join_quality",
            quality_passed,
            quality_details,
        )
    )

    pdf_text = _extract_text(
        root / "output" / "pdf" / "quantverse_v2_executive_research_report.pdf"
    )
    html_path = root / "output" / "html" / "quantverse_v2_research_report.html"
    html_text = (
        html_path.read_text(encoding="utf-8", errors="ignore")
        if html_path.exists()
        else ""
    )
    pdf_section = _extract_named_section(
        pdf_text,
        "2. Portfolio Holdings and Concentration",
        "3. Out-of-Sample Path Evidence",
    )
    html_section = _extract_named_section(
        html_text,
        '<h2 id="portfolio">Portfolio holdings</h2>',
        "<h2>Visual Portfolio Analytics</h2>",
    )
    pdf_lines = {line.strip().lower() for line in pdf_section.splitlines()}
    html_headers = {
        re.sub(r"<[^>]+>", "", header).strip().lower()
        for header in re.findall(r"<th[^>]*>(.*?)</th>", html_section, re.DOTALL)
    }
    checks.extend(
        [
            _check(
                "report_selected_stocks_uses_listing_country",
                "listing country" in pdf_lines and "listing_country" in html_headers,
                "checked current executive PDF and HTML holdings headers",
            ),
            _check(
                "report_selected_stocks_uses_issuer_country",
                "issuer country" in pdf_lines and "issuer_country" in html_headers,
                "checked current executive PDF and HTML holdings headers",
            ),
            _check(
                "report_no_ambiguous_country_header",
                "country" not in pdf_lines and "country" not in html_headers,
                f"pdf_country_header={'country' in pdf_lines}; html_country_header={'country' in html_headers}",
            ),
            _check(
                "report_no_ambiguous_currency_header",
                "currency" not in pdf_lines and "currency" not in html_headers,
                f"pdf_currency_header={'currency' in pdf_lines}; html_currency_header={'currency' in html_headers}",
            ),
        ]
    )

    economic_coverage = _economic_country_coverage(quality, exposure_quality)
    disclosure = (
        "Economic-country exposure is unavailable and is not inferred from "
        "listing venue, trading currency or issuer domicile."
    )
    disclosure_required = economic_coverage == 0.0
    disclosure_present = _normalize_report_text(disclosure) in _normalize_report_text(
        pdf_text
    ) and _normalize_report_text(disclosure) in _normalize_report_text(html_text)
    checks.append(
        _check(
            "report_economic_country_unavailable_disclosed",
            not disclosure_required or disclosure_present,
            f"economic_coverage={economic_coverage}; disclosure_present={disclosure_present}",
        )
    )

    if not view.empty and required_columns.issubset(view.columns):
        economic_values = view["economic_country"].fillna("").astype(str).str.lower()
        economic_not_inferred = bool(
            economic_coverage > 0.0
            or economic_values.isin({"unavailable", "missing", ""}).all()
        )
        foreign = view.loc[
            view["adr_or_foreign_issuer_flag"].map(
                lambda value: str(value).strip().lower() == "true"
            )
        ]
        foreign_preserved = bool(
            foreign.empty
            or foreign["listing_country"]
            .astype(str)
            .ne(foreign["issuer_country"].astype(str))
            .all()
        )
    else:
        economic_not_inferred = False
        foreign_preserved = False
    checks.append(
        _check(
            "selected_stock_report_view_no_economic_country_inference",
            economic_not_inferred,
            f"economic_coverage={economic_coverage}",
        )
    )
    checks.append(
        _check(
            "selected_stock_report_view_foreign_issuer_semantics_preserved",
            foreign_preserved,
            "flagged foreign issuers must retain distinct listing and issuer countries",
        )
    )

    excel_path = root / "output" / "excel" / "quantverse_v2_research_output.xlsx"
    try:
        excel_selected = _read_excel_sheet_table(
            excel_path, "SELECTED_STOCKS", header_row=3
        )
        excel_columns = set(excel_selected.columns.astype(str))
    except Exception as exc:
        excel_selected = pd.DataFrame()
        excel_columns = set()
        excel_error = _portable_exception_details(exc, excel_path)
    else:
        excel_error = ""
    excel_required = {"listing_country", "issuer_country", "economic_country"}
    checks.append(
        _check(
            "excel_selected_stocks_semantic_columns",
            excel_required.issubset(excel_columns)
            and len(excel_selected) == selected_count,
            f"columns={sorted(excel_columns)}; rows={len(excel_selected)}; error={excel_error}",
        )
    )
    checks.append(
        _check(
            "excel_selected_stocks_not_raw_legacy_table",
            bool(excel_columns)
            and "country" not in excel_columns
            and "currency" not in excel_columns,
            f"legacy_country={'country' in excel_columns}; legacy_currency={'currency' in excel_columns}",
        )
    )

    failed = [check for check in checks if not check["passed"]]
    return {
        "overall_status": "passed" if not failed else "failed",
        "failed_check_count": len(failed),
        "checks": checks,
    }


def _selected_score_count(scores: pd.DataFrame) -> int:
    if scores.empty:
        return 0
    if "selection_flag" not in scores:
        return len(scores)
    return int(
        scores["selection_flag"]
        .map(lambda value: str(value).strip().lower() in {"1", "true", "yes", "y"})
        .sum()
    )


def _economic_country_coverage(
    semantic_quality: pd.DataFrame,
    exposure_quality: pd.DataFrame,
) -> float:
    for frame in [semantic_quality, exposure_quality]:
        if not frame.empty and "economic_country_coverage_ratio" in frame:
            return _float(frame["economic_country_coverage_ratio"].iloc[0])
    return 0.0


def _extract_named_section(text: str, start: str, end: str) -> str:
    start_index = text.find(start)
    if start_index < 0:
        return ""
    end_index = text.find(end, start_index + len(start))
    return text[start_index : end_index if end_index >= 0 else len(text)]


def _normalize_report_text(text: str) -> str:
    return " ".join(text.lower().split())


def _check_current_v2_reports_no_stale_decisions(
    root: Path,
    summary: dict[str, object],
    checks: list[dict[str, object]],
) -> None:
    paths = [
        root / "output" / "pdf" / "quantverse_v2_research_report.pdf",
        root / "output" / "pdf" / "quantverse_v2_executive_research_report.pdf",
        root / "output" / "pdf" / "quantverse_v2_methodology_validation_appendix.pdf",
        root / "output" / "html" / "quantverse_v2_research_report.html",
        root / "output" / "excel" / "quantverse_v2_research_output.xlsx",
        root / "data" / "processed" / "quantverse_v2_demo_summary.json",
    ]
    text = "\n".join(_extract_text(path) for path in paths)
    hits = [
        pattern
        for pattern in STALE_CURRENT_DECISION_PATTERNS
        if re.search(pattern, text, flags=re.IGNORECASE)
    ]
    final_model = str(summary.get("final_selected_model", "")).strip()
    checks.append(
        _check(
            "current_v2_reports_no_stale_decision_phrases",
            not hits,
            f"final_model={final_model}; stale_hits={hits}",
        )
    )
    if final_model:
        checks.append(
            _check(
                "current_v2_reports_contain_final_model",
                final_model in text,
                f"final_model={final_model}",
            )
        )
    required_exposure_terms = [
        "Listing exposure",
        "Issuer exposure",
        "Economic exposure",
    ]
    missing_terms = [
        term for term in required_exposure_terms if term.lower() not in text.lower()
    ]
    checks.append(
        _check(
            "current_v2_reports_distinguish_exposure_types",
            not missing_terms,
            f"missing_terms={missing_terms}",
        )
    )


def _excel_sheet_names(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        return list(workbook.sheetnames)
    except ImportError:
        return _excel_sheet_names_from_zip(path)


def _excel_sheet_names_from_zip(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as archive:
        workbook_xml = archive.read("xl/workbook.xml")
    root = ElementTree.fromstring(workbook_xml)
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return [
        sheet.attrib["name"]
        for sheet in root.findall("main:sheets/main:sheet", namespace)
    ]


def _read_excel_sheet_table(
    path: Path,
    sheet_name: str,
    *,
    header_row: int,
) -> pd.DataFrame:
    """Read a small XLSX table without requiring an optional Excel engine."""

    with zipfile.ZipFile(path) as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        sheet = next(
            (
                item
                for item in workbook_root.findall(
                    f"{{{main_ns}}}sheets/{{{main_ns}}}sheet"
                )
                if item.attrib.get("name") == sheet_name
            ),
            None,
        )
        if sheet is None:
            raise ValueError(f"missing Excel sheet: {sheet_name}")
        relationship_id = sheet.attrib[f"{{{rel_ns}}}id"]

        relationships = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        package_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        target = next(
            item.attrib["Target"]
            for item in relationships.findall(f"{{{package_ns}}}Relationship")
            if item.attrib.get("Id") == relationship_id
        )
        sheet_path = target.lstrip("/")
        if not sheet_path.startswith("xl/"):
            sheet_path = f"xl/{sheet_path}"

        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = [
                "".join(node.text or "" for node in item.findall(f".//{{{main_ns}}}t"))
                for item in shared_root.findall(f"{{{main_ns}}}si")
            ]
        sheet_root = ElementTree.fromstring(archive.read(sheet_path))

    rows: dict[int, dict[int, object]] = {}
    for row in sheet_root.findall(f".//{{{main_ns}}}sheetData/{{{main_ns}}}row"):
        row_number = int(row.attrib.get("r", "0"))
        if row_number < header_row:
            continue
        cells: dict[int, object] = {}
        for cell in row.findall(f"{{{main_ns}}}c"):
            reference = cell.attrib.get("r", "A1")
            column_index = _excel_column_index(reference)
            cell_type = cell.attrib.get("t", "")
            value_node = cell.find(f"{{{main_ns}}}v")
            raw_value = value_node.text if value_node is not None else ""
            if cell_type == "s" and raw_value:
                value: object = shared_strings[int(raw_value)]
            elif cell_type == "inlineStr":
                value = "".join(
                    node.text or "" for node in cell.findall(f".//{{{main_ns}}}t")
                )
            elif cell_type == "b":
                value = raw_value == "1"
            else:
                value = raw_value
            cells[column_index] = value
        rows[row_number] = cells

    header_cells = rows.get(header_row, {})
    if not header_cells:
        return pd.DataFrame()
    max_column = max(header_cells)
    headers = [str(header_cells.get(index, "")) for index in range(max_column + 1)]
    values = [
        [row.get(index, "") for index in range(max_column + 1)]
        for row_number, row in sorted(rows.items())
        if row_number > header_row and any(str(value) for value in row.values())
    ]
    return pd.DataFrame(values, columns=headers)


def _excel_column_index(reference: str) -> int:
    letters = re.match(r"[A-Z]+", reference.upper())
    if letters is None:
        raise ValueError(f"invalid Excel cell reference: {reference}")
    index = 0
    for character in letters.group(0):
        index = index * 26 + (ord(character) - ord("A") + 1)
    return index - 1


def _extract_text(path: Path) -> str:
    if not path.exists():
        return ""
    if path.suffix.lower() == ".pdf":
        try:
            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages[:5])
        except Exception:
            return ""
    if path.suffix.lower() == ".xlsx":
        return _extract_xlsx_text(path)
    return path.read_text(encoding="utf-8", errors="ignore")


def _extract_xlsx_text(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        with zipfile.ZipFile(path) as archive:
            names = [
                name
                for name in archive.namelist()
                if name.startswith("xl/sharedStrings")
                or name.startswith("xl/worksheets/sheet")
            ]
            return "\n".join(
                archive.read(name).decode("utf-8", errors="ignore") for name in names
            )
    except Exception:
        return ""


def _read_json(path: Path) -> dict[str, object]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv(path: Path) -> pd.DataFrame:
    # Preserve source float identity for content-hash reconciliation.
    return (
        pd.read_csv(path, float_precision="round_trip")
        if path.exists()
        else pd.DataFrame()
    )


def _date_hashes_by_group(
    frame: pd.DataFrame,
    group_column: str,
) -> dict[object, str]:
    if frame.empty or not {group_column, "Date"}.issubset(frame.columns):
        return {}
    hashes: dict[object, str] = {}
    for key, group in frame.groupby(group_column, sort=True):
        dates = (
            pd.to_datetime(group["Date"], errors="coerce")
            .dropna()
            .drop_duplicates()
            .sort_values()
        )
        payload = "\n".join(dates.dt.strftime("%Y-%m-%d")).encode("utf-8")
        hashes[key] = f"dates-{hashlib.sha256(payload).hexdigest()[:24]}"
    return hashes


def _stable_frame_hash(frame: pd.DataFrame, columns: list[str]) -> str:
    if frame.empty or not set(columns).issubset(frame.columns):
        return "missing"
    normalized = frame[columns].copy()
    for column in columns:
        if "date" in column or column.endswith(("_start", "_end")):
            normalized[column] = pd.to_datetime(
                normalized[column],
                errors="coerce",
            ).dt.strftime("%Y-%m-%d")
        elif pd.api.types.is_numeric_dtype(normalized[column]):
            numeric = pd.to_numeric(normalized[column], errors="coerce")
            normalized[column] = numeric.map(
                lambda value: (f"{float(value):.12g}" if pd.notna(value) else "missing")
            )
        else:
            normalized[column] = normalized[column].fillna("").astype(str)
    normalized = normalized.sort_values(columns, kind="stable")
    payload = normalized.to_csv(index=False, lineterminator="\n").encode("utf-8")
    return f"frame-{hashlib.sha256(payload).hexdigest()[:24]}"


def _float(value: object) -> float:
    try:
        if value is None or pd.isna(value):
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _portable_exception_details(exc: Exception, path: Path | None = None) -> str:
    """Describe an artifact failure without leaking a local absolute path."""
    artifact = PureWindowsPath(str(path)).name if path is not None else "unavailable"
    return f"error_type={type(exc).__name__}; artifact={artifact}"


def _check(name: str, passed: bool, details: str) -> dict[str, object]:
    return {"check": name, "passed": bool(passed), "details": details}


if __name__ == "__main__":
    sys.exit(main())
